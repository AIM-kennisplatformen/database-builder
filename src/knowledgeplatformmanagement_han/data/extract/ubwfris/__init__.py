from contextlib import closing
from datetime import date, datetime
from string import ascii_letters
from typing import IO, Annotated, cast
from uuid import UUID

from annotated_types import Ge, Le
from loguru import logger
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.cell import Cell, MergedCell
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from pydantic import BaseModel, FiniteFloat, RootModel, StrictBool, StringConstraints, ValidationError
from unidecode import unidecode

from knowledgeplatformmanagement_han.data.dao.dataqualityissue import Dataqualityissue
from knowledgeplatformmanagement_han.data.dao.datasink_ubwfris import DatasinkUbwfris
from knowledgeplatformmanagement_han.data.model.compositionproject import CompositionProject
from knowledgeplatformmanagement_han.data.model.educationalproject import Educationalproject
from knowledgeplatformmanagement_han.data.model.hoursbooked import HoursBooked
from knowledgeplatformmanagement_han.data.model.hoursbudgeted import HoursBudgeted
from knowledgeplatformmanagement_han.data.model.hoursprojected import HoursProjected
from knowledgeplatformmanagement_han.data.model.hoursremaining import HoursRemaining
from knowledgeplatformmanagement_han.data.model.learningcommunity import Learningcommunity
from knowledgeplatformmanagement_han.data.model.namelike_id_employee import NamelikeIdEmployee
from knowledgeplatformmanagement_han.data.model.namelike_id_ubwcostcentre import NamelikeIdUbwcostcentre
from knowledgeplatformmanagement_han.data.model.namelike_name import NamelikeName
from knowledgeplatformmanagement_han.data.model.operationalproject import Operationalproject
from knowledgeplatformmanagement_han.data.model.participationinternal import Participationinternal
from knowledgeplatformmanagement_han.data.model.personubwfris import PersonUbwfris
from knowledgeplatformmanagement_han.data.model.projectlike import Projectstatus
from knowledgeplatformmanagement_han.data.model.projectlikes import Projectlikes
from knowledgeplatformmanagement_han.data.model.projectmanagement import Projectmanagement
from knowledgeplatformmanagement_han.data.model.provenant import Source
from knowledgeplatformmanagement_han.data.model.researchproject import Researchproject
from knowledgeplatformmanagement_han.data.model.school import School
from knowledgeplatformmanagement_han.data.model.strategicpartnership import Strategicpartnership
from knowledgeplatformmanagement_han.data.model.subproject import Subproject
from knowledgeplatformmanagement_han.data.model.unclearproject import Unclearproject
from knowledgeplatformmanagement_han.settings.timesheets import (
    IB630,
    REGEX_PROJECTID_1,
    RHA025A,
    IB630Withoutbooked,
    ProjecttypeToProjecttypeinfo,
    Projecttypeinfo,
    RowIB630,
    RowIB630Withoutbooked,
    RowRHA025A,
)

type CelltypeExtracted = str | int | float | date | None
type RowExtracted = dict[str, CelltypeExtracted]


class TimesheetsWorksheetUnavailableError(ValueError):
    def __init__(self, *, name_worksheet_timesheets: str) -> None:
        super().__init__(f"Worksheet '{name_worksheet_timesheets}' isn’t available in this workbook.")


class TimesheetsAlreadyloadedError(ValueError):
    def __init__(self, *, uuid: UUID) -> None:
        super().__init__(
            f"Not loading worksheet (UUID: {uuid!s}), since a worksheet with this UUID has been processed earlier.",
        )


class TimesheetsFormatInvalidError(ValueError):
    def __init__(self, *, format_worksheet: str) -> None:
        super().__init__(f"'{format_worksheet}' isn’t a valid worksheet format.")


class TimesheetsRowUnusableError(ValueError):
    def __init__(self, *, uuid: UUID, index_row: int) -> None:
        super().__init__(
            f"Worksheet (UUID: {uuid!s}) row {index_row} isn’t a usable timesheet row, since mandatory information is "
            "missing.",
        )


class ExportPowerbiTimesheet(BaseModel):
    billable: StrictBool
    date: date
    hours_type: Annotated[str, StringConstraints(min_length=1)]
    hours: FiniteFloat
    namelike_id_employee: NamelikeIdEmployee
    namelike_id_ubw: Annotated[
        str,
        StringConstraints(
            pattern=r"^[a-zA-ZÌ]{2,3}[0-9]{3,6}[\-]{0,1}[0-9]{0,3}$",
        ),
    ]
    namelike_name: Annotated[str, StringConstraints(min_length=1)]
    project_type_name_powerbi: Annotated[str, StringConstraints(min_length=1)]


ExportPowerbiTimesheets = RootModel[list[ExportPowerbiTimesheet]]


class ExportPowerbiPerson(BaseModel):
    employmentcontract_ftepercentage: Annotated[int, Ge(0), Le(100)] | None = None
    namelike_full: Annotated[str, StringConstraints(min_length=1)]
    namelike_id_employee: NamelikeIdEmployee
    namelike_id_ubwcostcentre: NamelikeIdUbwcostcentre | None = None


ExportPowerbiPersons = RootModel[list[ExportPowerbiPerson]]


# pylint: disable-next=too-few-public-methods,too-many-instance-attributes
class Ubwfris:
    def __init__(self, datasink: DatasinkUbwfris) -> None:
        self.datasink = datasink

    def delete_worksheet(self, *, uuid: UUID) -> None:
        del self.datasink.uuid_to_dataqualityissues[uuid]
        del self.datasink.uuid_to_persons_missing[uuid]
        del self.datasink.uuid_to_worksheet[uuid]

    def reset(self) -> None:
        self.datasink.uuid_to_dataqualityissues = {}
        self.datasink.uuid_to_worksheet = {}

    async def load_worksheet(
        self,
        *,
        format_worksheet: str,
        file_workbook: IO[bytes],
        name_worksheet: str,
        uuid: UUID,
    ) -> UUID:
        # TODO: Check for actual duplicates, e.g. using hash value of workbook.
        if uuid not in self.datasink.uuid_to_worksheet:
            logger.info(
                "Loading worksheet (UUID: {uuid!s}) '{name_worksheet}' with format '{format_worksheet}' ...",
                format_worksheet=format_worksheet,
                name_worksheet=name_worksheet,
                uuid=uuid,
            )
            name_column_to_index_column: type[IB630 | IB630Withoutbooked | RHA025A]
            if format_worksheet == "IB630":
                name_column_to_index_column = IB630
            elif format_worksheet == "IB630_without_booked":
                name_column_to_index_column = IB630Withoutbooked
            elif format_worksheet == "RHA025A":
                name_column_to_index_column = RHA025A
            else:
                raise TimesheetsFormatInvalidError(format_worksheet=format_worksheet)
            with closing(
                # Interestingly, the `filename` parameter of `load_workbook()` allows file-like objects. See
                # documentation.
                openpyxl_load_workbook(
                    filename=file_workbook,
                    data_only=True,
                    keep_links=False,
                    read_only=True,
                ),
            ) as workbook:
                if (
                    name_worksheet in workbook
                    and (_worksheet_timesheets := workbook[name_worksheet])
                    and isinstance(
                        _worksheet_timesheets,
                        ReadOnlyWorksheet,
                    )
                ):
                    self.datasink.uuid_to_worksheet[uuid] = _worksheet_timesheets
                else:
                    raise TimesheetsWorksheetUnavailableError(name_worksheet_timesheets=name_worksheet)
                self._read_worksheet(uuid=uuid, name_column_to_index_column=name_column_to_index_column)
        else:
            raise TimesheetsAlreadyloadedError(uuid=uuid)
        return uuid

    # The source code isn't overly complex, because the branches are brief and clear.
    def _extract_row(  # noqa: C901, PLR0911, PLR0912
        self,
        *,
        do_record_dataqualityissues: bool = False,
        index_row: int,
        name_column_to_index_column: type[IB630 | IB630Withoutbooked | RHA025A],
        row: tuple[Cell | MergedCell, ...],
        uuid: UUID,
    ) -> None | RowRHA025A | RowIB630 | RowIB630Withoutbooked:
        row_preprocessed: RowExtracted = {}
        for name_column, index_column in name_column_to_index_column.__members__.items():
            cellvalue = row[index_column].value
            if isinstance(cellvalue, datetime):
                row_preprocessed[name_column] = cellvalue.date()
            elif isinstance(cellvalue, str):
                row_preprocessed[name_column] = cellvalue.strip()
            elif isinstance(cellvalue, int | float | None):
                row_preprocessed[name_column] = cellvalue
            else:
                logger.error(
                    "Worksheet (ID: {uuid!s}) cell {cell} (column named '{column}') contains '{value}' of unknown "
                    "type '{type}'.",
                    cell=row[index_column].coordinate,
                    column=name_column,
                    value=cellvalue,
                    type=type(cellvalue),
                    uuid=uuid,
                )
                return None
            if row_preprocessed[name_column] == "":
                row_preprocessed[name_column] = None
        if not row_preprocessed["projectid"] or not row_preprocessed["naammedewerker"]:
            raise TimesheetsRowUnusableError(index_row=index_row, uuid=uuid)
        # Falls back if projecttypecode, registratiedatum are not in the data but can be derived or generated from other
        # data attributes.
        if name_column_to_index_column is not RHA025A:
            if not row_preprocessed.get("projecttypecode") and (projectid := row_preprocessed.get("projectid")):
                if (projecttype := REGEX_PROJECTID_1.fullmatch(str(projectid).strip())) is None:
                    logger.trace("`projectid` '{projectid}' does not match regex.", projectid=projectid)
                else:
                    # The mapping uses only the first two characters while the projectid can be 3 characters.
                    projecttypecode = projecttype[1][:2]
            if not row_preprocessed.get("registratiedatum") and isinstance(row_preprocessed.get("periode"), int):
                year, month = str(row_preprocessed["periode"])[:4], str(row_preprocessed["periode"])[4:]
                if month == "00":
                    month = "01"
                registratiedatum = (
                    datetime.strptime(f"{year}-{month}-1", "%Y-%m-%d")
                    .replace(
                        tzinfo=self.datasink.configurationtimesheets.ZONEINFO_UBWFRIS,
                    )
                    .date()
                )
        else:
            registratiedatum = cast(date, row_preprocessed[RHA025A.registratiedatum.name])
            projecttypecode = cast(str, row_preprocessed[RHA025A.projecttypecode.name])
        if name_column_to_index_column is RHA025A:
            try:
                return RowRHA025A(
                    deelproject_status_nl=Projectstatus[
                        cast(str, row_preprocessed[RHA025A.deelproject_status_nl.name])
                    ],
                    deelprojectid=cast(str, row_preprocessed[RHA025A.deelprojectid.name]),
                    deelprojectnaam=cast(str, row_preprocessed[RHA025A.deelprojectnaam.name]),
                    employmentcontract_ftepercentage=(
                        cast(int, _employmentcontract_ftepercentage)
                        if (
                            _employmentcontract_ftepercentage := row_preprocessed[
                                RHA025A.employmentcontract_ftepercentage.name
                            ]
                        )
                        else None
                    ),
                    geboekte_uren=cast(float, row_preprocessed[RHA025A.geboekte_uren.name]),
                    medewerkerid=cast(str, row_preprocessed[RHA025A.medewerkerid.name]),
                    naammedewerker=cast(str, row_preprocessed[RHA025A.naammedewerker.name]),
                    namelike_department=cast(str, row_preprocessed[RHA025A.namelike_department.name]),
                    namelike_id_school=cast(str, row_preprocessed[RHA025A.namelike_id_school.name]),
                    namelike_id_ubwcostcentre_employee=cast(
                        str,
                        row_preprocessed[RHA025A.namelike_id_ubwcostcentre_employee.name],
                    ),
                    namelike_id_ubwcostcentre_project=cast(
                        str,
                        row_preprocessed[RHA025A.namelike_id_ubwcostcentre_project.name],
                    ),
                    project_status_nl=Projectstatus[cast(str, row_preprocessed[RHA025A.project_status_nl.name])],
                    projectid=cast(str, row_preprocessed[RHA025A.projectid.name]),
                    projectmanager=cast(str, row_preprocessed[RHA025A.projectmanager.name]),
                    projectnaam=cast(str, row_preprocessed[RHA025A.projectnaam.name]),
                    projecttypecode=projecttypecode,
                    registratiedatum=registratiedatum,
                )
            except ValidationError as exception:
                if do_record_dataqualityissues:
                    self.datasink.uuid_to_dataqualityissues[uuid].extend(
                        Dataqualityissue(
                            index_row=index_row,
                            input=error["input"],
                            attributes=error["loc"],
                            message=error["msg"],
                            type=error["type"],
                        )
                        for error in exception.errors(include_url=False)
                    )
                return None
        elif name_column_to_index_column is IB630Withoutbooked:
            try:
                return RowIB630Withoutbooked(
                    begrote_uren=cast(float, row_preprocessed[IB630.begrote_uren.name]),
                    naammedewerker=cast(str, row_preprocessed[IB630Withoutbooked.naammedewerker.name]),
                    namelike_id_ubwcostcentre_project=cast(
                        str,
                        row_preprocessed[IB630Withoutbooked.namelike_id_ubwcostcentre_project.name],
                    ),
                    periode=cast(int, row_preprocessed[IB630Withoutbooked.periode.name]),
                    prognose_uren=cast(float, row_preprocessed[IB630Withoutbooked.prognose_uren.name]),
                    projectid=cast(str, row_preprocessed[IB630Withoutbooked.projectid.name]),
                    projectnaam=cast(str, row_preprocessed[IB630Withoutbooked.projectnaam.name]),
                    projecttypecode=projecttypecode,
                    registratiedatum=registratiedatum,
                    resterende_uren=cast(float, row_preprocessed[IB630Withoutbooked.resterende_uren.name]),
                )
            except ValidationError as exception:
                if do_record_dataqualityissues:
                    self.datasink.uuid_to_dataqualityissues[uuid].extend(
                        Dataqualityissue(
                            index_row=index_row,
                            input=error["input"],
                            attributes=error["loc"],
                            message=error["msg"],
                            type=error["type"],
                        )
                        for error in exception.errors(include_url=False)
                    )
                return None
        elif name_column_to_index_column is IB630:
            try:
                return RowIB630(
                    namelike_id_ubwcostcentre_project=cast(
                        str,
                        row_preprocessed[IB630.namelike_id_ubwcostcentre_project.name],
                    ),
                    begrote_uren=cast(float, row_preprocessed[IB630.begrote_uren.name]),
                    geboekte_uren=cast(float, row_preprocessed[IB630.geboekte_uren.name]),
                    naammedewerker=cast(str, row_preprocessed[IB630.naammedewerker.name]),
                    periode=cast(int, row_preprocessed[IB630.periode.name]),
                    prognose_uren=cast(float, row_preprocessed[IB630Withoutbooked.prognose_uren.name]),
                    projectid=cast(str, row_preprocessed[IB630.projectid.name]),
                    projectnaam=cast(str, row_preprocessed[IB630.projectnaam.name]),
                    projecttypecode=projecttypecode,
                    registratiedatum=registratiedatum,
                    resterende_uren=cast(float, row_preprocessed[IB630.resterende_uren.name]),
                )
            except ValidationError as exception:
                if do_record_dataqualityissues:
                    self.datasink.uuid_to_dataqualityissues[uuid].extend(
                        Dataqualityissue(
                            index_row=index_row,
                            input=error["input"],
                            attributes=error["loc"],
                            message=error["msg"],
                            type=error["type"],
                        )
                        for error in exception.errors(include_url=False)
                    )
                return None
        # This case should be impossible.
        raise AssertionError()

    def _generate_namelike_id_employee(self, namelike_first: str, namelike_last: str, uuid: UUID) -> str:
        # TODO: Solve this data quality issue within the data.
        namelike_id_employee = f"999-{namelike_first} {namelike_last}"
        address_email = Ubwfris._produce_address_email(namelike_first=namelike_first, namelike_last=namelike_last)
        self.datasink.id_to_personubwfris[namelike_id_employee] = PersonUbwfris(
            address_email=address_email,
            namelike_first=namelike_first,
            namelike_id_employee=namelike_id_employee,
            namelike_last=namelike_last,
        )
        self.datasink.uuid_to_persons_missing[uuid].append(namelike_id_employee)
        return namelike_id_employee

    @staticmethod
    def _produce_address_email(namelike_first: str, namelike_last: str) -> str:
        # TODO: Use sourced e-mail address rather than guessed one.
        return (
            "".join(letter for letter in unidecode(namelike_first) if letter in ascii_letters).lower()
            + "."
            + "".join(letter for letter in unidecode(namelike_last) if letter in ascii_letters).lower()
        ) + "@han.nl"

    # TODO: Don't find by name, but by true @key. However, the source data lacks this key.
    def _find_id_employee_by_name(self, *, name_person: str, uuid: UUID) -> str:
        namelike_first, namelike_last = self._parse_name_person(name_person=name_person)
        for namelike_id_employee, person in self.datasink.id_to_personubwfris.items():
            if person.namelike_first == namelike_first and person.namelike_last == namelike_last:
                return namelike_id_employee
        # Couldn't find the person in the mapping, so generate a new ID.
        return self._generate_namelike_id_employee(
            namelike_first=namelike_first,
            namelike_last=namelike_last,
            uuid=uuid,
        )

    def _add_projectmanagement(self, *, rowrha025a: RowRHA025A, uuid: UUID) -> None:
        # TODO: Don't add by name, but by true @key. However, the source data lacks this key.
        namelike_id_ubw_subproject = rowrha025a.deelprojectid
        namelike_id_employee_manager = self._find_id_employee_by_name(name_person=rowrha025a.projectmanager, uuid=uuid)
        if (projectmanager := self.datasink.id_to_personubwfris.get(namelike_id_employee_manager)) and (
            subproject := self.datasink.id_to_subproject.get(namelike_id_ubw_subproject)
        ):
            self.datasink.projectmanagements.append(
                Projectmanagement(
                    projectmanager=projectmanager.to_key(),
                    managedsubproject=subproject.to_key(),
                ),
            )

    def _add_person(self, rowrha025a: RowRHA025A) -> None:
        namelike_first, namelike_last = Ubwfris._parse_name_person(name_person=rowrha025a.naammedewerker)
        address_email = Ubwfris._produce_address_email(namelike_first=namelike_first, namelike_last=namelike_last)
        if rowrha025a.medewerkerid:
            self.datasink.id_to_personubwfris[rowrha025a.medewerkerid] = PersonUbwfris(
                address_email=address_email,
                employmentcontract_ftepercentage=rowrha025a.employmentcontract_ftepercentage,
                namelike_first=namelike_first,
                namelike_id_employee=rowrha025a.medewerkerid,
                namelike_id_ubwcostcentre=rowrha025a.namelike_id_ubwcostcentre_employee,
                namelike_last=namelike_last,
            )

    def _add_compositionproject(self, rowrha025a: RowRHA025A) -> None:
        if (overarchingproject := self.datasink.id_to_projectlike.get(rowrha025a.projectid)) and (
            subproject := self.datasink.id_to_subproject.get(
                rowrha025a.deelprojectid,
            )
        ):
            self.datasink.compositionprojects.append(
                CompositionProject(
                    overarchingproject=overarchingproject.to_key(),
                    projectpart=subproject.to_key(),
                ),
            )

    def _add_project(self, row: RowRHA025A | RowIB630 | RowIB630Withoutbooked) -> None:
        if row.projectid in self.datasink.id_to_projectlike:
            return
        projecttypeinfo = ProjecttypeToProjecttypeinfo[row.projecttypecode].value
        projectclassifier_status = row.project_status_nl if isinstance(row, RowRHA025A) else None
        # TODO: `registratiedatum` isn't the filing date but the (possibly future) date to which the record pertains.
        # For now, don't set `datetime_end_recorded`.
        namelike_name = NamelikeName(
            confidence=0.1,
            source=Source.ubwfris,
            value=row.projectnaam,
        )
        match projecttypeinfo.projectlike_subentity:
            # TODO: Set correct `date_event_end` and `date_event_start`.
            case "educationalproject":
                self.datasink.id_to_educationalproject[row.projectid] = Educationalproject(
                    date_event_end=row.registratiedatum,
                    date_event_start=row.registratiedatum,
                    namelike_id_ubw=row.projectid,
                    namelike_id_ubwcostcentre=row.namelike_id_ubwcostcentre_project,
                    namelike_name=namelike_name,
                    projectclassifier_financial=projecttypeinfo.projectclassifier_financial,
                    projectclassifier_status=projectclassifier_status,
                )
            case "learningcommunity":
                self.datasink.id_to_learningcommunity[row.projectid] = Learningcommunity(
                    date_event_end=row.registratiedatum,
                    date_event_start=row.registratiedatum,
                    namelike_id_ubw=row.projectid,
                    namelike_id_ubwcostcentre=row.namelike_id_ubwcostcentre_project,
                    namelike_name=namelike_name,
                    projectclassifier_financial=projecttypeinfo.projectclassifier_financial,
                    projectclassifier_status=projectclassifier_status,
                )
            case "operationalproject":
                self.datasink.id_to_operationalproject[row.projectid] = Operationalproject(
                    date_event_end=row.registratiedatum,
                    date_event_start=row.registratiedatum,
                    namelike_id_ubw=row.projectid,
                    namelike_id_ubwcostcentre=row.namelike_id_ubwcostcentre_project,
                    namelike_name=namelike_name,
                    projectclassifier_financial=projecttypeinfo.projectclassifier_financial,
                    projectclassifier_status=projectclassifier_status,
                )
            case "researchproject":
                self.datasink.id_to_researchproject[row.projectid] = Researchproject(
                    date_event_end=row.registratiedatum,
                    date_event_start=row.registratiedatum,
                    namelike_id_ubw=row.projectid,
                    namelike_id_ubwcostcentre=row.namelike_id_ubwcostcentre_project,
                    namelike_name=namelike_name,
                    projectclassifier_financial=projecttypeinfo.projectclassifier_financial,
                    projectclassifier_status=projectclassifier_status,
                )
            case "strategicpartnership":
                self.datasink.id_to_strategicpartnership[row.projectid] = Strategicpartnership(
                    date_event_end=row.registratiedatum,
                    date_event_start=row.registratiedatum,
                    namelike_id_ubw=row.projectid,
                    namelike_id_ubwcostcentre=row.namelike_id_ubwcostcentre_project,
                    namelike_name=namelike_name,
                    projectclassifier_financial=projecttypeinfo.projectclassifier_financial,
                    projectclassifier_status=projectclassifier_status,
                )
            case _:
                self.datasink.id_to_unclearproject[row.projectid] = Unclearproject(
                    date_event_end=row.registratiedatum,
                    date_event_start=row.registratiedatum,
                    namelike_id_ubw=row.projectid,
                    namelike_id_ubwcostcentre=row.namelike_id_ubwcostcentre_project,
                    namelike_name=namelike_name,
                    projectclassifier_financial=projecttypeinfo.projectclassifier_financial,
                    projectclassifier_status=projectclassifier_status,
                )

    def _add_subproject(self, *, rowrha025a: RowRHA025A) -> None:
        if rowrha025a.deelprojectid in self.datasink.id_to_subproject:
            logger.trace("Not adding Subproject (ID: {}), as it was already added.", rowrha025a.deelprojectid)
            return
        subprojectclassifier_status = rowrha025a.deelproject_status_nl
        projecttypecode = rowrha025a.projecttypecode
        projecttype = ProjecttypeToProjecttypeinfo[projecttypecode].name
        projecttypeprojectinfo = ProjecttypeToProjecttypeinfo[projecttype].value
        namelike_name = NamelikeName(
            confidence=0.1,
            source=Source.ubwfris,
            value=rowrha025a.deelprojectnaam,
        )
        subproject = Subproject(
            namelike_id_ubwcostcentre=rowrha025a.namelike_id_ubwcostcentre_project,
            date_event_end=rowrha025a.registratiedatum,
            date_event_start=rowrha025a.registratiedatum,
            namelike_id_ubw=rowrha025a.deelprojectid,
            namelike_name=namelike_name,
            projectclassifier_financial=projecttypeprojectinfo.projectclassifier_financial,
            projectclassifier_status=subprojectclassifier_status,
        )
        self.datasink.id_to_subproject[rowrha025a.deelprojectid] = subproject

    def _add_participationinternals(self, rowrha025a: RowRHA025A) -> None:
        if (projectlike := self.datasink.id_to_projectlike.get(rowrha025a.projectid)) and (
            school := self.datasink.id_to_school[rowrha025a.namelike_id_school]
        ):
            internallycooperatingcomponents = projectlike.to_key()
            leadingcomponent = school.to_key()
            namelike_id_ubwcostcentre_employee = rowrha025a.namelike_id_ubwcostcentre_employee
            # TODO: Explain the conditional logic below.
            if (
                not namelike_id_ubwcostcentre_employee
                or str(namelike_id_ubwcostcentre_employee).strip()[:4] == rowrha025a.namelike_id_school.strip()[:4]
            ):
                participationinternal = Participationinternal(
                    internallycooperatingcomponents=internallycooperatingcomponents,
                    leadingcomponent=leadingcomponent,
                )
            else:
                id_school = f"{str(namelike_id_ubwcostcentre_employee).strip()[:4]}00"
                partnercomponent = (
                    self.datasink.id_to_school[id_school].to_key() if id_school in self.datasink.id_to_school else None
                )
                participationinternal = Participationinternal(
                    internallycooperatingcomponents=internallycooperatingcomponents,
                    leadingcomponent=leadingcomponent,
                    partnercomponent=partnercomponent,
                )
            self.datasink.participationinternals.append(participationinternal)

    def _add_hours(self, row: RowIB630 | RowIB630Withoutbooked | RowRHA025A, uuid: UUID) -> None:
        if isinstance(row, RowRHA025A):
            namelike_id_employee = row.medewerkerid
        else:
            namelike_id_employee = self._find_id_employee_by_name(
                name_person=row.naammedewerker,
                uuid=uuid,
            )
        person_employee = self.datasink.id_to_personubwfris[namelike_id_employee]
        person_hours = person_employee.to_key()
        projectlike_charges: Projectlikes | None
        if isinstance(row, RowRHA025A):
            namelike_id_ubw = row.deelprojectid
            if not (projectlike_charges := self.datasink.id_to_subproject.get(namelike_id_ubw)):
                return
        else:
            namelike_id_ubw = row.projectid
            if not (projectlike_charges := self.datasink.id_to_projectlike.get(namelike_id_ubw)):
                return
        charges_hours = projectlike_charges.to_key()
        projectclassifier_financial_to_projecttype = {
            enumvalue.value.projectclassifier_financial: enumname
            for enumname, enumvalue in ProjecttypeToProjecttypeinfo.__members__.items()
            if enumvalue.value.projectclassifier_financial
        }
        projecttype = projectclassifier_financial_to_projecttype[projectlike_charges.projectclassifier_financial]
        projecttypeinfo: Projecttypeinfo = ProjecttypeToProjecttypeinfo[projecttype].value
        billable: bool = projecttypeinfo.billable or (
            projectlike_charges.namelike_id_ubwcostcentre is not None
            and person_employee.namelike_id_ubwcostcentre is not None
            and (
                str(projectlike_charges.namelike_id_ubwcostcentre)[:4]
                != str(person_employee.namelike_id_ubwcostcentre)[:4]
            )
        )
        if isinstance(row, RowIB630Withoutbooked):
            self.datasink.timesheets.append(
                HoursBudgeted(
                    billable=billable,
                    budgets_hours=person_hours,
                    charges_hours=charges_hours,
                    date_event_registration=row.registratiedatum,
                    timesheets_hours=row.begrote_uren,
                ),
            )
            self.datasink.timesheets.append(
                HoursProjected(
                    billable=billable,
                    charges_hours=charges_hours,
                    date_event_registration=row.registratiedatum,
                    projects_hours=person_hours,
                    timesheets_hours=row.prognose_uren,
                ),
            )
            self.datasink.timesheets.append(
                HoursRemaining(
                    billable=billable,
                    charges_hours=charges_hours,
                    date_event_registration=row.registratiedatum,
                    timesheets_hours=row.resterende_uren,
                    remains_hours=person_hours,
                ),
            )
        if isinstance(row, RowIB630 | RowRHA025A):
            self.datasink.timesheets.append(
                HoursBooked(
                    billable=billable,
                    books_hours=person_hours,
                    charges_hours=charges_hours,
                    date_event_registration=row.registratiedatum,
                    timesheets_hours=row.geboekte_uren,
                ),
            )

    def _add_school(self, rowrha025a: RowRHA025A) -> None:
        if rowrha025a.namelike_id_school not in self.datasink.id_to_school:
            self.datasink.id_to_school[rowrha025a.namelike_id_school] = School(
                namelike_id_school=rowrha025a.namelike_id_school,
                namelike_department=rowrha025a.namelike_department,
            )

    def _process_row(
        self,
        *,
        row: RowRHA025A | RowIB630 | RowIB630Withoutbooked,
        uuid: UUID,
    ) -> None:
        if isinstance(row, RowRHA025A):
            self._add_subproject(rowrha025a=row)
        self._add_project(row=row)
        if isinstance(row, RowRHA025A):
            self._add_projectmanagement(rowrha025a=row, uuid=uuid)
            self._add_compositionproject(rowrha025a=row)
        self._add_hours(row=row, uuid=uuid)
        if isinstance(row, RowRHA025A):
            self._add_participationinternals(rowrha025a=row)

    # I don't see a way to simplify this.
    # pylint: disable-next=too-many-locals
    def _read_worksheet(  # noqa: C901, PLR0912
        self,
        *,
        name_column_to_index_column: type[IB630 | IB630Withoutbooked | RHA025A],
        uuid: UUID,
    ) -> None:
        # openpyxl is 1-based, but this is incremented immediately.
        header_row = 0
        # Ignore Mypy fault, because of indirectly related issue https://github.com/python/mypy/issues/17184.
        self.datasink.uuid_to_dataqualityissues[uuid] = []
        self.datasink.uuid_to_persons_missing[uuid] = []

        for index_row, row in enumerate(
            iterable=self.datasink.uuid_to_worksheet[uuid].iter_rows(min_row=header_row + 1),  # type: ignore[misc]
            start=header_row + 1,
        ):
            header_row += 1
            # This would theoretically fail if the column names are just falsy numeric values that incorrectly aren't
            # typed as string even though they are header names, or empty strings. However, our header names are never
            # like that.
            if any(True for value in row if value.value):
                logger.debug(
                    "Worksheet’s (UUID: {uuid!s}) header is on row {index_row}. Every next row will be read.",
                    index_row=index_row,
                    uuid=uuid,
                )
                break
        if name_column_to_index_column is RHA025A:
            ## First all persons & schools need to be added as they are used in relationships.
            # Ignore Mypy fault, because of indirectly related issue https://github.com/python/mypy/issues/17184.
            for index_row, row in enumerate(
                iterable=self.datasink.uuid_to_worksheet[uuid].iter_rows(min_row=header_row + 1),  # type: ignore[misc]
                start=header_row + 1,
            ):
                try:
                    rowrha025a = self._extract_row(
                        index_row=index_row,
                        name_column_to_index_column=name_column_to_index_column,
                        row=row,
                        uuid=uuid,
                    )
                except TimesheetsRowUnusableError as exception:
                    # TODO: Handle along with Pydantic validation and record data quality issue.
                    logger.error("{}", exception)
                else:
                    if isinstance(rowrha025a, RowRHA025A):
                        self._add_person(rowrha025a=rowrha025a)
                        self._add_school(rowrha025a=rowrha025a)

        ## Then the other entities can be added, we need the the employee id in hourallocation and projectmanagement.
        # Ignore Mypy fault, because of indirectly related issue https://github.com/python/mypy/issues/17184.
        for index_row, row in enumerate(
            iterable=self.datasink.uuid_to_worksheet[uuid].iter_rows(min_row=header_row + 1),  # type: ignore[misc]
            start=header_row + 1,
        ):
            try:
                row_extracted = self._extract_row(
                    index_row=index_row,
                    name_column_to_index_column=name_column_to_index_column,
                    row=row,
                    uuid=uuid,
                    do_record_dataqualityissues=True,
                )
            except TimesheetsRowUnusableError:
                # Already reported in previous run.
                pass
            else:
                if row_extracted is not None:
                    self._process_row(row=row_extracted, uuid=uuid)
        if self.datasink.uuid_to_dataqualityissues.get(uuid):
            logger.info("Data quality issue(s) found in worksheet (UUID: {!s}).", uuid)
        else:
            logger.info("No data quality issues found in worksheet (UUID: {!s}).", uuid)
        if self.datasink.uuid_to_persons_missing.get(uuid):
            logger.info("Missing Person(s) in worksheet (UUID: {!s}).", uuid)
        else:
            logger.info("No missing Persons in worksheet (UUID: {!s}).", uuid)
        if "projecttypecode" not in name_column_to_index_column.__members__:
            logger.warning(
                "`projecttypecode` isn’t mapped in the worksheet format `{}` for worksheet (UUID: {!s}). The type of "
                "project can be derived from the `projectid`, but this derivation isn’t always correct.",
                name_column_to_index_column.__name__,
                uuid,
            )

    @staticmethod
    def _parse_name_person(name_person: str) -> tuple[str, str]:
        if len(splitted := name_person.split(maxsplit=1)) == 2:  # noqa: PLR2004
            name_first, name_last = splitted
        else:
            logger.error("Person name {name} cannot be split.", name=name_person)
            name_first, name_last = (name_person, name_person)
        return name_first, name_last
